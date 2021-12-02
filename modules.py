import os
import glob
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
import settings


# 罫線を引く処理
def draw_boader (sheet, start_row, end_row, start_col, end_col) :
    # フォルダ名/ファイル名が記入されている列より左側か右側かを判定するためのフラグの値
    LEFT_CELL_FLG   = 0
    RIGHT_CELL_FLG  = 1

    # 罫線のフォーマット
    b_top_left  = Border(top=Side(style='thin', color='000000'),    left=Side(style='thin', color='000000') )
    b_top       = Border(top=Side(style='thin', color='000000')                                             )
    b_left      = Border(                                           left=Side(style='thin', color='000000') )

    for row in range(start_row, end_row + 1):
        cell_flag = LEFT_CELL_FLG   # フォルダ名/ファイル名が記入されている列より左側か右側かを判定するためのフラグ

        for col in range(start_col, end_col + 1):
            # セルを取得
            cell = sheet.cell(row=row, column=col)

            # セルが空 & その行のフォルダ名/ファイル名が記載されている列よりも左側の列
            if cell.value == None and cell_flag == LEFT_CELL_FLG:
                cell.border = b_left
            
            # セルが空でない
            elif cell.value != None :
                cell.border = b_top_left
                cell_flag = RIGHT_CELL_FLG

            # セルが空 & その行のフォルダ名/ファイル名が記載されている列よりも右側の列
            elif cell.value == None and cell_flag == RIGHT_CELL_FLG:
                cell.border = b_top

# OSに対応したファイルパスのデリミタを返す関数
def get_delimiter () :
    if os.name == "nt":         # Windows
        return "¥"

    elif os.name == 'posix':    # Mac or Linux
        return "/"


# 除外リストに含まれているフォルダ名/ファイル名が書き出し対象に含まれているかチェックする処理(完全一致)
def check_exclusion_file (path, delimiter, exclusion_list) :
    path = path.split(delimiter)
    is_exclusion_file = (set(path) & set(exclusion_list)) != set()

    return is_exclusion_file

# ファイル一覧を書き出す処理
def write_file_list (sheet, start_row, start_col, delimiter) :
    row = start_row
    col = start_col

    # 比較用のリストを作成
    pre_file = []

    # 同階層以下のフォルダ/ファイル一覧を取得
    f_list = glob.glob("**", recursive=True)

    # 一覧を書き出す処理
    for file in f_list:

        # 除外リストに含まれているフォルダ名/ファイル名の場合は書き出しせずに終了
        is_exclusion_file = check_exclusion_file(file, delimiter, settings.exclusion_list)
        if is_exclusion_file == True :
            pass
        else :
            file = file.split(delimiter)    #取得した一覧をデリミタで分割し、リスト化

            for i, f_name in enumerate(file):
                if len(pre_file) == 0 or i >= len(pre_file) or file[i] != pre_file[i]:
                    sheet.cell(row=row, column=col).value = f_name

                col += 1    # 次の列に移動

            pre_file = file

            row += 1        # 次の行に移動
            col = start_col # 列の位置を初期位置に移動

# 背景色を白にする処理
def set_background_color (sheet) :
    for row in sheet:
        for cell in row:
            sheet[cell.coordinate].fill = PatternFill(patternType='solid', fgColor='ffffff')

# 列幅調整
def set_column_width (sheet) :
    for col in range(1, sheet.max_column) :
        col_letter = sheet.cell(row=1,column=col).column_letter
        sheet.column_dimensions[col_letter].width = settings.col_widht

    col_letter = sheet.cell(row=1,column=sheet.max_column).column_letter
    sheet.column_dimensions[col_letter].width = settings.end_col_widht