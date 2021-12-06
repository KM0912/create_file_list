import settings
import openpyxl
import modules

# エクセルファイルの作成
book_title  = settings.book_title   # ファイル名
sheet_title = settings.sheet_title  # シート名
book        = openpyxl.Workbook()   # 新規Bookの作成
sheet       = book.worksheets[0]    # シート設定
sheet.title = sheet_title           # シートのタイトル設定

# セル開始位置
start_row   = settings.start_row  # 行の初期位置
start_col   = settings.start_col  # 列の初期位置

# OSを判定し、ファイルパスの区切り文字を決定
delimiter = modules.get_delimiter()

# 一覧を書き出す処理
modules.write_file_list(sheet, start_row, start_col, delimiter)

# 背景色の設定
modules.set_background_color(sheet)

# 罫線を引く処理
modules.draw_boader(sheet, start_row, sheet.max_row, start_col, sheet.max_column)

# 列幅調整
modules.set_column_width(sheet)

# 保存 & 終了
book.save(book_title)
book.close()